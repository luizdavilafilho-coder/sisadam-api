import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import pathlib
import os
import cv2 
import openpyxl # Biblioteca instalada para gerar o .xlsx nativo
from openpyxl.styles import Font, PatternFill # Para pintar o cabeçalho do Excel
from PIL import Image
from datetime import datetime

# --- CONFIGURAÇÃO DE CAMINHOS ---
PASTA_BASE = pathlib.Path.home() / "Documents" / "SISADAM"
PASTA_BANCO = PASTA_BASE / "Arquivos_Banco"
PASTA_FOTOS = PASTA_BASE / "Fotos_Moradores"
ARQUIVO_DB_MORADORES = PASTA_BANCO / "dados_moradores.db"

class BancoMoradores(ctk.CTkFrame):
    def __init__(self, master, nivel="OPERADOR", **kwargs):
        super().__init__(master, **kwargs)
        
        self.cor_destaque = "#F57C00"
        self.nivel = nivel.strip().upper()
        self.id_selecionado = None
        self.caminho_foto_atual = None
        self.img_tk = None
        
        PASTA_BANCO.mkdir(parents=True, exist_ok=True)
        PASTA_FOTOS.mkdir(parents=True, exist_ok=True)
        
        self.inicializar_banco()
        
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        # --- SEÇÃO 0: TOPO (DADOS E FOTO) ---
        self.f_topo = ctk.CTkFrame(self, fg_color="transparent")
        self.f_topo.grid(row=0, column=0, sticky="nsew", padx=10, pady=(10, 5))
        self.f_topo.grid_columnconfigure(0, weight=1)
        self.f_topo.grid_columnconfigure(1, weight=0)

        # 1. ÁREA DE DADOS
        self.f_dados = ctk.CTkFrame(self.f_topo, fg_color="#121212", corner_radius=10, border_width=1, border_color="#333")
        self.f_dados.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        estilo_in = {"fg_color": "white", "text_color": "black", "placeholder_text_color": "gray"}
        
        ctk.CTkLabel(self.f_dados, text="IDENTIFICAÇÃO", text_color=self.cor_destaque, font=("Arial", 12, "bold")).grid(row=0, column=0, columnspan=4, pady=(10, 5))
        self.cb_bloco = ctk.CTkComboBox(self.f_dados, values=["BLOCO A", "BLOCO B"], width=100, fg_color="white", text_color="black")
        self.cb_bloco.set("BLOCO A"); self.cb_bloco.grid(row=1, column=0, padx=10, pady=5, sticky="w")
        self.ent_apto = ctk.CTkEntry(self.f_dados, placeholder_text="Apto", width=80, **estilo_in)
        self.ent_apto.grid(row=1, column=1, padx=5, pady=5, sticky="w")
        self.ent_nome = ctk.CTkEntry(self.f_dados, placeholder_text="Nome do Morador", width=400, **estilo_in)
        self.ent_nome.grid(row=1, column=2, columnspan=2, padx=5, pady=5, sticky="w")
        
        self.ent_ddi = ctk.CTkEntry(self.f_dados, placeholder_text="55", width=50, **estilo_in)
        self.ent_ddi.insert(0, "55"); self.ent_ddi.grid(row=2, column=0, padx=10, pady=5, sticky="w")
        self.ent_ddd = ctk.CTkEntry(self.f_dados, placeholder_text="11", width=50, **estilo_in)
        self.ent_ddd.insert(0, "11"); self.ent_ddd.grid(row=2, column=1, padx=5, pady=5, sticky="w")
        self.ent_fone = ctk.CTkEntry(self.f_dados, placeholder_text="Fone", width=150, **estilo_in)
        self.ent_fone.grid(row=2, column=2, padx=5, pady=5, sticky="w")
        self.ent_email = ctk.CTkEntry(self.f_dados, placeholder_text="E-mail", width=240, **estilo_in)
        self.ent_email.grid(row=2, column=3, padx=5, pady=5, sticky="w")

        ctk.CTkLabel(self.f_dados, text="EMERGÊNCIA", text_color="#2196F3", font=("Arial", 12, "bold")).grid(row=3, column=0, columnspan=4, pady=(15, 5))
        
        self.ent_contato1 = ctk.CTkEntry(self.f_dados, placeholder_text="Contato 1", width=180, **estilo_in)
        self.ent_contato1.grid(row=4, column=0, columnspan=2, padx=10, pady=5, sticky="w")
        self.ent_fone1 = ctk.CTkEntry(self.f_dados, placeholder_text="Fone 1", width=150, **estilo_in)
        self.ent_fone1.grid(row=4, column=2, padx=5, pady=5, sticky="w")
        self.ent_grau1 = ctk.CTkEntry(self.f_dados, placeholder_text="Grau", width=100, **estilo_in)
        self.ent_grau1.grid(row=4, column=3, padx=5, pady=5, sticky="w")

        self.ent_contato2 = ctk.CTkEntry(self.f_dados, placeholder_text="Contato 2", width=180, **estilo_in)
        self.ent_contato2.grid(row=5, column=0, columnspan=2, padx=10, pady=5, sticky="w")
        self.ent_fone2 = ctk.CTkEntry(self.f_dados, placeholder_text="Fone 2", width=150, **estilo_in)
        self.ent_fone2.grid(row=5, column=2, padx=5, pady=5, sticky="w")
        self.ent_grau2 = ctk.CTkEntry(self.f_dados, placeholder_text="Grau", width=100, **estilo_in)
        self.ent_grau2.grid(row=5, column=3, padx=5, pady=5, sticky="w")

        self.ent_obs = ctk.CTkEntry(self.f_dados, placeholder_text="OBSERVAÇÕES MÉDICAS", width=680, **estilo_in)
        self.ent_obs.grid(row=6, column=0, columnspan=4, padx=10, pady=15, sticky="w")

        # --- BARRA DE AÇÕES ---
        self.f_acoes = ctk.CTkFrame(self.f_dados, fg_color="transparent")
        self.f_acoes.grid(row=7, column=0, columnspan=4, padx=10, pady=(0, 10), sticky="ew")
        
        self.var_lgpd = ctk.StringVar(value="off")
        self.chk_lgpd = ctk.CTkCheckBox(self.f_acoes, text="Consentimento LGPD", variable=self.var_lgpd, onvalue="on", offvalue="off", text_color="#2196F3")
        self.chk_lgpd.pack(side="left", padx=5)
        
        self.btn_novo = ctk.CTkButton(self.f_acoes, text="+ NOVO REGISTRO", fg_color="#2E7D32", width=120, command=self.preparar_novo)
        self.btn_novo.pack(side="left", padx=(10, 5))

        self.btn_gravar = ctk.CTkButton(self.f_acoes, text="GRAVAR", fg_color=self.cor_destaque, text_color="black", width=100, command=self.salvar_registro)
        self.btn_gravar.pack(side="left", padx=5)

        self.btn_editar = ctk.CTkButton(self.f_acoes, text="EDITAR", fg_color="#1E88E5", width=100, command=self.habilitar_edicao)
        self.btn_editar.pack(side="left", padx=5)

        ctk.CTkButton(self.f_acoes, text="📸 FOTO", fg_color="#5D4037", width=90, command=self.capturar_foto).pack(side="left", padx=5)
        
        self.btn_recarregar = ctk.CTkButton(self.f_acoes, text="🔄 CANCELAR / RECARREGAR", fg_color="#555", width=150, command=self.reset_modulo)
        self.btn_recarregar.pack(side="left", padx=5)
        
        # --- NOVOS BOTÕES EXCLUSIVOS MASTER ---
        self.btn_deletar = ctk.CTkButton(self.f_acoes, text="EXCLUIR REGISTRO", fg_color="#D32F2F", width=140, command=self.deletar_registro)
        self.btn_deletar.pack(side="right", padx=5)

        self.btn_cofre = ctk.CTkButton(self.f_acoes, text="ENVIAR PARA COFRE", fg_color="#607D8B", width=160, command=self.enviar_para_cofre)
        self.btn_cofre.pack(side="right", padx=5)

        # 2. ÁREA DA FOTO
        self.f_foto = ctk.CTkFrame(self.f_topo, width=240, height=340, fg_color="#121212", corner_radius=10, border_width=1, border_color=self.cor_destaque)
        self.f_foto.grid(row=0, column=1, sticky="ns")
        self.f_foto.grid_propagate(False)
        ctk.CTkLabel(self.f_foto, text="IDENTIFICAÇÃO VISUAL", text_color=self.cor_destaque, font=("Arial", 10, "bold")).pack(pady=5)
        self.lbl_preview = ctk.CTkLabel(self.f_foto, text="SEM FOTO", width=200, height=220, fg_color="#1E1E1E", corner_radius=8)
        self.lbl_preview.pack(pady=5, padx=20)

        # --- SEÇÃO 1: BUSCA INTELIGENTE ---
        self.f_busca = ctk.CTkFrame(self, fg_color="#1a1a1a", height=50, corner_radius=0)
        self.f_busca.grid(row=1, column=0, sticky="ew", padx=10, pady=5)
        ctk.CTkLabel(self.f_busca, text="BUSCAR POR:", font=("Arial", 11, "bold")).pack(side="left", padx=(15, 5))
        self.cb_filtro = ctk.CTkComboBox(self.f_busca, values=["NOME", "APTO", "BLOCO"], width=100)
        self.cb_filtro.set("NOME"); self.cb_filtro.pack(side="left", padx=5)
        self.ent_busca = ctk.CTkEntry(self.f_busca, placeholder_text="Digite para filtrar...", width=400)
        self.ent_busca.pack(side="left", padx=10, pady=10)
        self.ent_busca.bind("<KeyRelease>", lambda e: self.carregar_dados(self.ent_busca.get()))
        ctk.CTkButton(self.f_busca, text="LIMPAR BUSCA", fg_color="#555", width=120, command=self.limpar_busca).pack(side="left", padx=5)

        # --- SEÇÃO 2: TREEVIEW COMPLETO ---
        self.f_tabela = ctk.CTkFrame(self, fg_color="transparent")
        self.f_tabela.grid(row=2, column=0, sticky="nsew", padx=10, pady=(0, 10))
        self.f_tabela.grid_rowconfigure(0, weight=1)
        self.f_tabela.grid_columnconfigure(0, weight=1)

        cols = ("ID", "BLOCO", "APTO", "NOME", "TELEFONE", "E-MAIL", "CONTATO 1", "FONE 1", "GRAU 1", "CONTATO 2", "FONE 2", "GRAU 2", "OBSERVAÇÕES", "REGISTRO")
        self.tree = ttk.Treeview(self.f_tabela, columns=cols, show="headings")
        
        sc_y = ttk.Scrollbar(self.f_tabela, orient="vertical", command=self.tree.yview)
        sc_x = ttk.Scrollbar(self.f_tabela, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=sc_y.set, xscrollcommand=sc_x.set)

        larguras = [40, 70, 60, 200, 130, 180, 150, 120, 80, 150, 120, 80, 250, 130]
        for i, c in enumerate(cols):
            self.tree.heading(c, text=c)
            self.tree.column(c, width=larguras[i], anchor="center" if i not in [3, 5, 12] else "w")
            
        self.tree.grid(row=0, column=0, sticky="nsew")
        sc_y.grid(row=0, column=1, sticky="ns")
        sc_x.grid(row=1, column=0, sticky="ew")
        
        self.tree.bind("<<TreeviewSelect>>", self.ao_selecionar)
        
        self.aplicar_permissoes()
        self.carregar_dados()
        self.bloquear_campos_padrao()

    def inicializar_banco(self):
        conn = sqlite3.connect(str(ARQUIVO_DB_MORADORES))
        conn.execute("CREATE TABLE IF NOT EXISTS moradores (id INTEGER PRIMARY KEY AUTOINCREMENT)")
        colunas = [("bloco", "TEXT"), ("apartamento", "TEXT"), ("nome_morador", "TEXT"), ("ddi", "TEXT"), ("ddd", "TEXT"), 
                   ("telefone", "TEXT"), ("email", "TEXT"), ("contato1", "TEXT"), ("fone1", "TEXT"), ("grau1", "TEXT"),
                   ("contato2", "TEXT"), ("fone2", "TEXT"), ("grau2", "TEXT"),
                   ("obs_medicas", "TEXT"), ("lgpd", "TEXT"), ("data_registro", "TEXT"), ("foto", "TEXT")]
        for col, tipo in colunas:
            try: conn.execute(f"ALTER TABLE moradores ADD COLUMN {col} {tipo}")
            except: pass
        conn.commit(); conn.close()

    def aplicar_permissoes(self):
        if self.nivel != "MASTER":
            self.btn_deletar.pack_forget()
            self.btn_cofre.pack_forget()

    # REGRA CRÍTICA DE LGPD / ACESSO
    def aplicar_mascara(self, tipo, valor):
        if not valor or str(valor).strip() in ["", "None", "-"]: return ""
        
        if self.nivel in ["ADMIN", "MASTER"]: 
            return str(valor)
            
        if self.nivel == "OPERADOR":
            if tipo in ["email", "emergencia", "obs"]: 
                return "***"
        return str(valor)

    def reset_modulo(self):
        self.limpar_campos(reset_id=True)
        self.bloquear_campos_padrao()
        self.carregar_dados()
        self.cb_filtro.set("NOME")
        self.ent_busca.delete(0, 'end')

    def preparar_novo(self):
        self.limpar_campos(reset_id=True)
        
        campos_base = [self.ent_apto, self.ent_nome, self.ent_ddi, self.ent_ddd, self.ent_fone]
        for e in campos_base:
            try: e.configure(state="normal")
            except: pass
        try: self.cb_bloco.configure(state="normal")
        except: pass

        if self.nivel in ["ADMIN", "MASTER"]:
            campos_admin = [self.ent_email, self.ent_contato1, self.ent_fone1, self.ent_grau1, self.ent_contato2, self.ent_fone2, self.ent_grau2, self.ent_obs]
            for e in campos_admin:
                try: e.configure(state="normal")
                except: pass
            try: self.chk_lgpd.configure(state="normal")
            except: pass

        try: self.btn_gravar.configure(state="normal", text="GRAVAR NOVO")
        except: pass

    def bloquear_campos_padrao(self):
        campos_entrada = [self.ent_apto, self.ent_nome, self.ent_ddi, self.ent_ddd, self.ent_fone, self.ent_email, 
                          self.ent_contato1, self.ent_fone1, self.ent_grau1, self.ent_contato2, self.ent_fone2, self.ent_grau2, self.ent_obs]
        for e in campos_entrada:
            try: e.configure(state="disabled")
            except: pass
        try: self.cb_bloco.configure(state="disabled")
        except: pass
        try: self.chk_lgpd.configure(state="disabled")
        except: pass
        try: self.btn_gravar.configure(state="disabled")
        except: pass

    def habilitar_edicao(self):
        if not self.id_selecionado:
            messagebox.showwarning("Aviso", "Selecione um morador na tabela para editar.")
            return

        campos_base = [self.ent_apto, self.ent_nome, self.ent_ddi, self.ent_ddd, self.ent_fone]
        for e in campos_base:
            try: e.configure(state="normal")
            except: pass
        try: self.cb_bloco.configure(state="normal")
        except: pass

        if self.nivel in ["ADMIN", "MASTER"]:
            campos_admin = [self.ent_email, self.ent_contato1, self.ent_fone1, self.ent_grau1, self.ent_contato2, self.ent_fone2, self.ent_grau2, self.ent_obs]
            for e in campos_admin:
                try: e.configure(state="normal")
                except: pass
            try: self.chk_lgpd.configure(state="normal")
            except: pass

        try: self.btn_gravar.configure(state="normal", text="ATUALIZAR DADOS")
        except: pass

    def ao_selecionar(self, event):
        sel = self.tree.selection()
        if not sel: return
        self.id_selecionado = self.tree.item(sel[0])['values'][0]
        
        try:
            conn = sqlite3.connect(str(ARQUIVO_DB_MORADORES))
            conn.row_factory = sqlite3.Row 
            d = conn.execute("SELECT * FROM moradores WHERE id=?", (self.id_selecionado,)).fetchone()
            conn.close()
        except: return
            
        if d:
            campos = [self.ent_apto, self.ent_nome, self.ent_ddi, self.ent_ddd, self.ent_fone, self.ent_email, self.ent_contato1, self.ent_fone1, self.ent_grau1, self.ent_contato2, self.ent_fone2, self.ent_grau2, self.ent_obs]
            for e in campos:
                try: e.configure(state="normal")
                except: pass
            try: self.cb_bloco.configure(state="normal")
            except: pass
            try: self.chk_lgpd.configure(state="normal")
            except: pass

            def injetar_texto(widget, valor):
                try:
                    widget.delete(0, 'end')
                    if valor and str(valor).strip() not in ["", "None"]:
                        widget.insert(0, str(valor))
                except: pass

            def buscar_dado(coluna, padrao=""):
                try:
                    val = d[coluna]
                    return str(val) if val is not None and str(val).strip() not in ["", "None"] else padrao
                except: return padrao

            try: self.cb_bloco.set(buscar_dado("bloco", "BLOCO A"))
            except: pass
            
            injetar_texto(self.ent_apto, buscar_dado("apartamento"))
            injetar_texto(self.ent_nome, buscar_dado("nome_morador"))
            injetar_texto(self.ent_ddi, buscar_dado("ddi", "55"))
            injetar_texto(self.ent_ddd, buscar_dado("ddd", "11"))
            injetar_texto(self.ent_fone, buscar_dado("telefone"))
            
            injetar_texto(self.ent_email, self.aplicar_mascara("email", buscar_dado("email")))
            injetar_texto(self.ent_contato1, self.aplicar_mascara("emergencia", buscar_dado("contato1")))
            injetar_texto(self.ent_fone1, self.aplicar_mascara("emergencia", buscar_dado("fone1")))
            injetar_texto(self.ent_grau1, self.aplicar_mascara("emergencia", buscar_dado("grau1")))
            injetar_texto(self.ent_contato2, self.aplicar_mascara("emergencia", buscar_dado("contato2")))
            injetar_texto(self.ent_fone2, self.aplicar_mascara("emergencia", buscar_dado("fone2")))
            injetar_texto(self.ent_grau2, self.aplicar_mascara("emergencia", buscar_dado("grau2")))
            injetar_texto(self.ent_obs, self.aplicar_mascara("obs", buscar_dado("obs_medicas")))

            try: self.var_lgpd.set(buscar_dado("lgpd", "off"))
            except: pass
            
            try:
                foto_bd = buscar_dado("foto")
                if foto_bd: self.caminho_foto_atual = os.path.normpath(foto_bd)
                else: self.caminho_foto_atual = None
            except: self.caminho_foto_atual = None
                
            self.atualizar_preview()
            self.bloquear_campos_padrao()

    def atualizar_preview(self):
        if self.caminho_foto_atual and os.path.exists(self.caminho_foto_atual):
            try:
                with Image.open(self.caminho_foto_atual) as img_pil:
                    img_pil.thumbnail((200, 220))
                    img_memoria = img_pil.copy()
                self.img_tk = ctk.CTkImage(light_image=img_memoria, dark_image=img_memoria, size=(img_memoria.width, img_memoria.height))
                self.lbl_preview.configure(image=self.img_tk, text="")
                self.lbl_preview.image = self.img_tk 
            except Exception:
                self.lbl_preview.configure(image=None, text="ERRO NA FOTO")
                self.lbl_preview.image = None
        else:
            self.lbl_preview.configure(image=None, text="SEM FOTO")
            self.lbl_preview.image = None

    def carregar_dados(self, filtro=""):
        for i in self.tree.get_children(): self.tree.delete(i)
        conn = sqlite3.connect(str(ARQUIVO_DB_MORADORES))
        query = "SELECT id, bloco, apartamento, nome_morador, ('+'||ddi||' ('||ddd||') '||telefone), email, contato1, fone1, grau1, contato2, fone2, grau2, obs_medicas, data_registro FROM moradores"
        
        if filtro:
            campo = {"NOME": "nome_morador", "APTO": "apartamento", "BLOCO": "bloco"}[self.cb_filtro.get()]
            query += f" WHERE {campo} LIKE ?"
            cur = conn.execute(query + " ORDER BY nome_morador ASC", (f"%{filtro.upper()}%",))
        else:
            cur = conn.execute(query + " ORDER BY nome_morador ASC")
            
        for r in cur.fetchall():
            linha = list(r)
            linha[5] = self.aplicar_mascara("email", linha[5])
            linha[6] = self.aplicar_mascara("emergencia", linha[6])
            linha[7] = self.aplicar_mascara("emergencia", linha[7])
            linha[8] = self.aplicar_mascara("emergencia", linha[8])
            linha[9] = self.aplicar_mascara("emergencia", linha[9])
            linha[10] = self.aplicar_mascara("emergencia", linha[10])
            linha[11] = self.aplicar_mascara("emergencia", linha[11])
            linha[12] = self.aplicar_mascara("obs", linha[12])
            
            self.tree.insert("", "end", values=linha)
        conn.close()

    def salvar_registro(self):
        b, apto, nome = self.cb_bloco.get(), self.ent_apto.get().strip(), self.ent_nome.get().upper()
        
        if not nome or not apto: 
            messagebox.showwarning("Atenção", "NOME e APTO são obrigatórios.")
            return
            
        caminho_para_salvar = self.caminho_foto_atual if self.caminho_foto_atual else ""
        conn = sqlite3.connect(str(ARQUIVO_DB_MORADORES))
        conn.row_factory = sqlite3.Row

        if self.id_selecionado:
            old = conn.execute("SELECT * FROM moradores WHERE id=?", (self.id_selecionado,)).fetchone()
            
            email = old["email"] if self.nivel == "OPERADOR" else self.ent_email.get()
            contato1 = old["contato1"] if self.nivel == "OPERADOR" else self.ent_contato1.get()
            fone1 = old["fone1"] if self.nivel == "OPERADOR" else self.ent_fone1.get()
            grau1 = old["grau1"] if self.nivel == "OPERADOR" else self.ent_grau1.get()
            contato2 = old["contato2"] if self.nivel == "OPERADOR" else self.ent_contato2.get()
            fone2 = old["fone2"] if self.nivel == "OPERADOR" else self.ent_fone2.get()
            grau2 = old["grau2"] if self.nivel == "OPERADOR" else self.ent_grau2.get()
            obs = old["obs_medicas"] if self.nivel == "OPERADOR" else self.ent_obs.get()
            lgpd = old["lgpd"] if self.nivel == "OPERADOR" else self.var_lgpd.get()
            
            if not caminho_para_salvar and old["foto"] and str(old["foto"]) != "None":
                caminho_para_salvar = old["foto"]
                
            dados = (b, apto, nome, self.ent_ddi.get(), self.ent_ddd.get(), self.ent_fone.get(), email,
                     contato1, fone1, grau1, contato2, fone2, grau2, obs, lgpd, datetime.now().strftime('%d/%m/%Y %H:%M'), caminho_para_salvar)
            
            conn.execute("""UPDATE moradores SET bloco=?, apartamento=?, nome_morador=?, ddi=?, ddd=?, telefone=?, email=?,
                            contato1=?, fone1=?, grau1=?, contato2=?, fone2=?, grau2=?, obs_medicas=?, lgpd=?, data_registro=?, foto=? WHERE id=?""", (*dados, self.id_selecionado))
            messagebox.showinfo("Sucesso", "Cadastro atualizado com segurança.")
        else:
            email = "" if self.nivel == "OPERADOR" else self.ent_email.get()
            contato1 = "" if self.nivel == "OPERADOR" else self.ent_contato1.get()
            fone1 = "" if self.nivel == "OPERADOR" else self.ent_fone1.get()
            grau1 = "" if self.nivel == "OPERADOR" else self.ent_grau1.get()
            contato2 = "" if self.nivel == "OPERADOR" else self.ent_contato2.get()
            fone2 = "" if self.nivel == "OPERADOR" else self.ent_fone2.get()
            grau2 = "" if self.nivel == "OPERADOR" else self.ent_grau2.get()
            obs = "" if self.nivel == "OPERADOR" else self.ent_obs.get()
            lgpd = "off" if self.nivel == "OPERADOR" else self.var_lgpd.get()
            
            dados = (b, apto, nome, self.ent_ddi.get(), self.ent_ddd.get(), self.ent_fone.get(), email,
                     contato1, fone1, grau1, contato2, fone2, grau2, obs, lgpd, datetime.now().strftime('%d/%m/%Y %H:%M'), caminho_para_salvar)
            
            conn.execute("""INSERT INTO moradores (bloco, apartamento, nome_morador, ddi, ddd, telefone, email, 
                            contato1, fone1, grau1, contato2, fone2, grau2, obs_medicas, lgpd, data_registro, foto) 
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", dados)
            messagebox.showinfo("Sucesso", "Novo morador cadastrado.")
            
        conn.commit(); conn.close()
        self.reset_modulo()

    def limpar_campos(self, reset_id=True):
        campos = [self.ent_apto, self.ent_nome, self.ent_ddi, self.ent_ddd, self.ent_fone, self.ent_email, self.ent_contato1, self.ent_fone1, self.ent_grau1, self.ent_contato2, self.ent_fone2, self.ent_grau2, self.ent_obs]
        for e in campos:
            try: e.configure(state="normal")
            except: pass
        try: self.cb_bloco.configure(state="normal")
        except: pass
        try: self.chk_lgpd.configure(state="normal")
        except: pass

        if reset_id: 
            self.id_selecionado = None
            try: self.btn_gravar.configure(state="normal", text="GRAVAR")
            except: pass
            
        for e in campos:
            try: e.delete(0, 'end')
            except: pass
            
        if reset_id:
            try:
                self.ent_ddi.insert(0, "55")
                self.ent_ddd.insert(0, "11")
            except: pass
            
        try: self.var_lgpd.set("off")
        except: pass
        
        self.caminho_foto_atual = None
        try:
            self.lbl_preview.configure(image=None, text="SEM FOTO")
            self.lbl_preview.image = None
        except: pass

    def limpar_busca(self):
        self.ent_busca.delete(0, 'end'); self.carregar_dados()

    def deletar_registro(self):
        if self.nivel != "MASTER" or not self.id_selecionado: return
        if messagebox.askyesno("Excluir", "Remover morador definitivamente? Esta ação não pode ser desfeita."):
            conn = sqlite3.connect(str(ARQUIVO_DB_MORADORES))
            conn.execute("DELETE FROM moradores WHERE id=?", (self.id_selecionado,))
            conn.commit(); conn.close()
            self.reset_modulo()

    # --- FUNÇÃO DO COFRE COM EXCEL NATIVO (.XLSX) ---
    def enviar_para_cofre(self):
        if messagebox.askyesno("Cofre de Segurança", "Deseja exportar TODOS os registros do banco de dados para uma planilha Excel nativa (.xlsx)?"):
            try:
                pasta_cofre = PASTA_BASE / "Cofre_Backup"
                pasta_cofre.mkdir(parents=True, exist_ok=True)
                
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                arquivo_excel = pasta_cofre / f"cofre_moradores_{timestamp}.xlsx"
                
                conn = sqlite3.connect(str(ARQUIVO_DB_MORADORES))
                cur = conn.cursor()
                cur.execute("SELECT * FROM moradores")
                
                colunas = [desc[0].upper() for desc in cur.description]
                linhas = cur.fetchall()
                conn.close()
                
                # Cria a planilha nativa do Excel
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Cofre SISADAM"
                
                # Adiciona e formata o cabeçalho (Fundo Laranja, Texto Branco, Negrito)
                ws.append(colunas)
                for celula in ws[1]:
                    celula.font = Font(bold=True, color="FFFFFF")
                    celula.fill = PatternFill(start_color="F57C00", end_color="F57C00", fill_type="solid")
                
                # Adiciona os dados reais das linhas
                for linha in linhas:
                    ws.append(list(linha))
                    
                wb.save(arquivo_excel)
                messagebox.showinfo("Cofre - Exportação Concluída", f"Todos os registros foram salvos nativamente no Excel:\n\n{arquivo_excel}")
                
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao gerar planilha do cofre. Verifique se o openpyxl está instalado. Erro detalhado: {e}")

    def capturar_foto(self):
        cap = cv2.VideoCapture(0)
        while True:
            ret, frame = cap.read()
            cv2.imshow("Captura SISADAM (ESPACO=Foto | ESC=Sair)", frame)
            k = cv2.waitKey(1)
            if k%256 == 27: break
            elif k%256 == 32:
                path = PASTA_FOTOS / f"morador_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
                cv2.imwrite(str(path), frame)
                self.caminho_foto_atual = os.path.normpath(str(path))
                self.atualizar_preview(); break
        cap.release(); cv2.destroyAllWindows()